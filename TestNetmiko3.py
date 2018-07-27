import openpyxl
import re
from netmiko import ConnectHandler
from collections import OrderedDict
from itertools import combinations


# Make Spreadsheet
wb = openpyxl.Workbook()
ws1 = wb.create_sheet('Device')
ws2 = wb.create_sheet('Mem_CPU')
ws3 = wb.create_sheet('Buffer')
ws4 = wb.create_sheet('Summary')

#Make Formatting
alignment = openpyxl.styles.Alignment(
    horizontal='center', vertical='center', wrap_text=True)
font = openpyxl.styles.Font(bold=True)
#Remove Default ExtraSheet
extraSheet = wb.get_sheet_by_name('Sheet')
wb.remove_sheet(extraSheet)

#Write Header1
ws1.cell(row=1, column=1).value = 'Hostname'
ws1.cell(row=1, column=2).value = 'PID'
ws1.cell(row=1, column=3).value = 'Description'
ws1.cell(row=1, column=4).value = 'SN'
ws1.cell(row=1, column=5).value = 'Version'
ws1.cell(row=1, column=6).value = 'IOS'
ws1.cell(row=1, column=7).value = 'Uptime'
ws1.cell(row=1, column=8).value = 'DRAM'
ws1.cell(row=1, column=9).value = 'FLASH'
for cell in ws1["1:1"]:
    cell.font = font
    cell.alignment = alignment
#Write Header2
ws2.cell(row=1, column=2).value = 'Memory'
ws2.merge_cells('B1:D1')
ws2.cell(row=2, column=1).value = 'Hostname'
ws2.cell(row=2, column=2).value = 'Total'
ws2.cell(row=2, column=3).value = 'Used'
ws2.cell(row=2, column=4).value = 'Free'
ws2.merge_cells('E1:F1')
ws2.cell(row=2, column=5).value = 'Memory Utilization'
ws2.cell(row=2, column=6).value = 'Recomendation'
ws2.merge_cells('G1:I1')
ws2.cell(row=1, column=7).value = 'CPU'
ws2.cell(row=2, column=7).value = 'five seconds'
ws2.cell(row=2, column=8).value = 'one minute'
ws2.cell(row=2, column=9).value = 'five minute'
ws2.cell(row=2, column=10).value = 'recomendation'
for cell in ws2["1:1"]:
    cell.font = font
    cell.alignment = alignment
for cell in ws2["2:2"]:
    cell.font = font
    cell.alignment = alignment
#Write Header3
ws3.cell(row=1, column=1).value = 'Hostname'
ws3.cell(row=1, column=2).value = 'Variable'
ws3.cell(row=1, column=3).value = 'hits'
ws3.cell(row=1, column=4).value = 'misses'
ws3.cell(row=1, column=5).value = 'Total'
ws3.cell(row=1, column=6).value = 'Percent'
ws3.cell(row=1, column=7).value = 'Recomendation'
for cell in ws3["1:1"]:
    cell.font = font
    cell.alignment = alignment


#Connect to Devices
wb2 = openpyxl.load_workbook('DeviceList.xlsx', data_only=True)
rs = wb2["DeviceList"]

mylist = []

for cols in rs.iter_cols(min_row=1, min_col=1):
    for cell in cols:
        mylist.append(cell.value)

ip_index = mylist.index("ip")
user_index = mylist.index("username")
pass_index = mylist.index("password")
mylistIP = mylist[ip_index+1:user_index]
mylistUser = mylist[user_index+1:pass_index]
mylistPass = mylist[pass_index+1:]

# cisco_ios1 = {
#     'device_type': 'cisco_ios',
#     'ip':   '10.10.10.5',
#     'username': 'poweruser',
#     'password': 'admin'
# }
#
# cisco_ios2 = {
#     'device_type': 'cisco_ios',
#     'ip':   '10.10.10.6',
#     'username': 'poweruser2',
#     'password': 'admin'
# }
#
# cisco_ios3 = {
#     'device_type': 'cisco_ios',
#     'ip':   '10.10.10.7',
#     'username': 'poweruser3',
#     'password': 'admin'
# }

# all_devices = [cisco_ios1, cisco_ios2, cisco_ios3]


#loop through devices
for ip, user, passw in zip(mylistIP, mylistUser, mylistPass):
    net_connect = ConnectHandler(device_type='cisco_ios', ip=ip, username=user, password=passw)
    # command
    output = net_connect.send_command('show running-config')
    output = output + net_connect.send_command('show inventory')
    output = output + net_connect.send_command('show version')
    output = output + net_connect.send_command('show memory statistics')
    output = output + net_connect.send_command('show processes cpu sorted')
    output = output + net_connect.send_command('show buffers')

    #Parse
    hostname = re.findall(r'^hostname (.*)\s*', output, re.I | re.M)
    PID = re.findall(r'\s*PID:(.*),\sVID', output, re.I | re.M)
    Description = re.findall(r'\sDESCR:\s*(.*)\s*', output, re.I | re.M)
    SN = re.findall(r'\s*SN:\s*(.*)\s*', output, re.I | re.M)
    IOS = re.findall(r'\((.*)\),\sVersion', output, re.I | re.M)
    Version = re.findall(r'\sVersion\s(.*),', output, re.I | re.M)
    Uptime = re.findall(r'\s*uptime\sis\s(.*)', output, re.I | re.M)
    DRAM = re.findall(r'\swith\s(.*)\/', output, re.I | re.M)
    MemTot = re.findall(r'^Processor\s+\w+\s+(\d+)\s+', output, re.I|re.M)
    MemUse = re.findall(r'^Processor\s+\w+\s+\d+\s+(\d+)\s+', output, re.I|re.M)
    MemFree = re.findall(r'^Processor\s+\w+\s+\d+\s+\d+\s+(\d+)\s+', output, re.I|re.M)
    Cpu5sec = re.findall(r'\sfive\sseconds:\s(.*);\sone', output, re.I|re.M)
    Cpu1min = re.findall(r'\sone\sminute:\s(.*);\sfive', output, re.I|re.M)
    Cpu5min = re.findall(r'\sfive\sminutes:\s(.*)\s*', output, re.I|re.M)
    smallhits = re.findall(r'^Small.*\s+.*\s+(\d+)\s+hits,', output, re.I|re.M)
    smallmiss = re.findall(r'^Small.*\s+.*\s+.*(\d+)\s+misses,', output, re.I|re.M)
    middlehits = re.findall(r'^Middle.*\s+.*\s+(\d+)\s+hits,', output, re.I|re.M)
    middlemiss = re.findall(r'^Middle.*\s+.*\s+.*(\d+)\s+misses,', output, re.I|re.M)
    bighits = re.findall(r'^Big.*\s+.*\s+(\d+)\s+hits,', output, re.I|re.M)
    bigmiss = re.findall(r'^Big.*\s+.*\s+.*(\d+)\s+misses,', output, re.I|re.M)
    verybighits = re.findall(r'^VeryBig.*\s+.*\s+(\d+)\s+hits,', output, re.I|re.M)
    verybigmiss = re.findall(r'^VeryBig.*\s+.*\s+.*(\d+)\s+misses,', output, re.I|re.M)
    largehits = re.findall(r'^Large.*\s+.*\s+(\d+)\s+hits,', output, re.I|re.M)
    largemiss = re.findall(r'^Large.*\s+.*\s+.*(\d+)\s+misses,', output, re.I|re.M)
    hugehits = re.findall(r'^Huge.*\s+.*\s+(\d+)\s+hits,', output, re.I|re.M)
    hugemiss = re.findall(r'^Huge.*\s+.*\s+.*(\d+)\s+misses,', output, re.I|re.M)


    #sheet Device
    dictdev = OrderedDict(
        [('HostDict', hostname), ('PIDdict', PID), ('DescDict', Description), ('SNdict', SN), ('VersionDict', Version),
         ('IOSdict', IOS), ('UptimeDict', Uptime), ('DRAMdict', DRAM)])

    row1 = ws1.max_row + 1
    col1 = 1

    for key in dictdev.keys():
        i = 0
        for item in dictdev[key]:
            if len(item) > 1:
                ws1.cell(row=row1 + i, column=col1).value = item
                i += 1
            else:
                ws1.cell(row=row1, column=col1).value = item
        col1 += 1

        # count = ws1.max_row
        # print(count)


    #Sheet mem_cpu
    dictmem = OrderedDict([('MemTotList', MemTot), ('MemUsedList', MemUse), ('MemFreeList', MemFree)])
    dictcpu = OrderedDict([('Cpu5secList', Cpu5sec), ('Cpu1minList', Cpu1min), ('Cpu5minList', Cpu5min)])
    dicthost = dict(HostList=hostname)


    row2 = ws2.max_row + 1
    col2 = 1

    for key in dicthost.keys():
        for item in dicthost[key]:
            ws2.cell(row=row2, column=col2).value = item
        col2 += 1

    for key in dictmem.keys():
        for item in dictmem[key]:
            ws2.cell(row=row2, column=col2).value = int(item)
        col2 += 1


    for key in dictcpu.keys():
        y = 2
        i = 0
        for item in dictcpu[key]:
            if len(item) > 1:
                ws2.cell(row=row2+i, column=col2+y).value = item
                # ws2.cell(row=row + i, column=col + y).number_format = '0'
                i += 1
            else:
                ws2.cell(row=row2, column=col2+y).value = item
                # ws2.cell(row=row, column=col + y).number_format = '0'
        col2 += 1
    ws2["E3"] = "=IF(ISERROR(C3/B3),0,(C3/B3)*100)"
    ws2["F3"] = """=IF((VALUE(I3)*100)<=40,"Excellent",IF((VALUE(I3)*100)<=60,"Good",IF((VALUE(I3)*100)<=80,"Fair","Poor")))"""

    #Sheet Buffer
    dicthost = dict(HostList=hostname)
    dictSmall = OrderedDict([('SmallHits', smallhits), ('SmallMisses', smallmiss)])
    dictMiddle = OrderedDict([('MiddleHits', middlehits), ('MiddleMisses', middlemiss)])
    dictBig = OrderedDict([('BigHits', bighits), ('BigMisses', bigmiss)])
    dictVeryBig = OrderedDict([('VeryBigHits', verybighits), ('VeryBigMisses', verybigmiss)])
    dictLarge = OrderedDict([('LargeHits', largehits), ('LargeMisses', largemiss)])
    dictHuge = OrderedDict([('HugeHits', hugehits), ('HugeMisses', hugemiss)])

    dictCollectedBuff = OrderedDict([('Small', dictSmall), ('Middle', dictMiddle),
                                     ('Big', dictBig), ('VeryBig', dictVeryBig), ('Large', dictLarge),
                                     ('Huge', dictHuge)])

    row3 = ws3.max_row + 1
    col3 = 1
    i = 0
    j = 0

    for key in dicthost.keys():
        for item in dicthost[key]:
            ws3.cell(row=row3, column=col3).value = item
            col3 += 1

    for key in dictCollectedBuff.keys():
        col3 = 2
        ws3.cell(row=row3 + i, column=col3).value = key
        col3 += 1
        for key2 in dictCollectedBuff[key]:
            for item in dictCollectedBuff[key][key2]:
                ws3.cell(row=row3 + i, column=col3).value = int(item)
                col3 += 1
        i += 1
    ws3["E2"] = "=C2+D2"
    ws3["F2"] = "=IF(ISERROR(D2/E2),0,(D2/E2)*100)"
    ws3["G2"] = """=IF(F2<=5,"Excellent",IF(F2<=10,"Good",IF(F2<=20,"Fair","Poor")))"""
    print(hostname)

wb.save('sample.xlsx')

# net_connect = ConnectHandler(**cisco_ios2)
#
# #command
# output = net_connect.send_command('show running-config')
# output = output + net_connect.send_command('show inventory')
# output = output + net_connect.send_command('show version')
# output = output + net_connect.send_command('show memory statistics')
# output = output + net_connect.send_command('show processes cpu sorted')
# output = output + net_connect.send_command('show buffers')

