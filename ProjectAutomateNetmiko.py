import os
import openpyxl
from openpyxl.chart import (
    Reference,
    BarChart3D,
)
import re
import sys
from time import sleep
from netmiko import ConnectHandler
from collections import OrderedDict
import xlwings as xw


#Define Counter
cannotCount = 0

# Make Spreadsheet
wb = openpyxl.Workbook()
ws1 = wb.create_sheet('Device')
ws2 = wb.create_sheet('Mem_CPU')
ws3 = wb.create_sheet('Buffer')
ws4 = wb.create_sheet('Summary')

#Make Formatting
alignment = openpyxl.styles.Alignment(
    horizontal='center', vertical='center', wrap_text=True)
font = openpyxl.styles.Font(bold=True)
#Remove Default ExtraSheet
extraSheet = wb.get_sheet_by_name('Sheet')
wb.remove_sheet(extraSheet)

def MakeSpreadshet():
    #Write Header1
    ws1.cell(row=1, column=1).value = 'Hostname'
    ws1.cell(row=1, column=2).value = 'PID'
    ws1.cell(row=1, column=3).value = 'Description'
    ws1.cell(row=1, column=4).value = 'SN'
    ws1.cell(row=1, column=5).value = 'Version'
    ws1.cell(row=1, column=6).value = 'IOS'
    ws1.cell(row=1, column=7).value = 'Uptime'
    ws1.cell(row=1, column=8).value = 'DRAM'
    ws1.cell(row=1, column=9).value = 'FLASH'
    for cell in ws1["1:1"]:
        cell.font = font
        cell.alignment = alignment
    #Write Header2
    ws2.cell(row=1, column=2).value = 'Memory'
    ws2.merge_cells('B1:D1')
    ws2.cell(row=2, column=1).value = 'Hostname'
    ws2.cell(row=2, column=2).value = 'Total'
    ws2.cell(row=2, column=3).value = 'Used'
    ws2.cell(row=2, column=4).value = 'Free'
    ws2.merge_cells('E1:F1')
    ws2.cell(row=2, column=5).value = 'Memory Utilization'
    ws2.cell(row=2, column=6).value = 'Recomendation'
    ws2.merge_cells('G1:I1')
    ws2.cell(row=1, column=7).value = 'CPU'
    ws2.cell(row=2, column=7).value = 'five seconds'
    ws2.cell(row=2, column=8).value = 'one minute'
    ws2.cell(row=2, column=9).value = 'five minute'
    ws2.cell(row=2, column=10).value = 'recomendation'
    for cell in ws2["1:1"]:
        cell.font = font
        cell.alignment = alignment
    for cell in ws2["2:2"]:
        cell.font = font
        cell.alignment = alignment
    #Write Header3
    ws3.cell(row=1, column=1).value = 'Hostname'
    ws3.cell(row=1, column=2).value = 'Variable'
    ws3.cell(row=1, column=3).value = 'hits'
    ws3.cell(row=1, column=4).value = 'misses'
    ws3.cell(row=1, column=5).value = 'Total'
    ws3.cell(row=1, column=6).value = 'Percent'
    ws3.cell(row=1, column=7).value = 'Recomendation'
    for cell in ws3["1:1"]:
        cell.font = font
        cell.alignment = alignment
    #Write Header4
    ws4.cell(row=1, column=1).value = 'Hostname'
    ws4.cell(row=1, column=2).value = 'Memory'
    ws4.cell(row=1, column=3).value = 'CPU'
    ws4.cell(row=1, column=4).value = 'Buffers'
    ws4.cell(row=1, column=5).value = 'Conclusion'
    for cell in ws4["1:1"]:
        cell.font = font
        cell.alignment = alignment


def SheetDevice():
    # sheet Device
    dictdev = OrderedDict(
        [('HostDict', hostname), ('PIDdict', PID), ('DescDict', Description), ('SNdict', SN), ('VersionDict', Version),
         ('IOSdict', IOS), ('UptimeDict', Uptime), ('DRAMdict', DRAM)])

    row = ws1.max_row + 1
    col = 1

    for key in dictdev.keys():
        i = 0
        for item in dictdev[key]:
            if len(item) > 1:
                ws1.cell(row=row + i, column=col).value = item
                i += 1
            else:
                ws1.cell(row=row, column=col).value = item
        col += 1

def SheetMemCPU():
    # Sheet mem_cpu
    dictmem = OrderedDict([('MemTotList', MemTot), ('MemUsedList', MemUse), ('MemFreeList', MemFree)])
    dictcpu = OrderedDict([('Cpu5secList', Cpu5sec), ('Cpu1minList', Cpu1min), ('Cpu5minList', Cpu5min)])
    dicthost = dict(HostList=hostname)

    row = ws2.max_row + 1
    col = 1

    for key in dicthost.keys():
        for item in dicthost[key]:
            ws2.cell(row=row, column=col).value = item
        col += 1

    for key in dictmem.keys():
        for item in dictmem[key]:
            ws2.cell(row=row, column=col).value = int(item)
        col += 1

    for key in dictcpu.keys():
        y = 2
        i = 0
        for item in dictcpu[key]:
            if len(item) > 1:
                ws2.cell(row=row + i, column=col + y).value = item
                i += 1
            else:
                ws2.cell(row=row, column=col + y).value = item
        col += 1

    for row, cellObj in enumerate(list(ws2.columns)[4], start=1):
        if row == 1:
            continue
        elif row == 2:
            continue
        else:
            n = '=IF(ISERROR(C%d/B%d),0,(C%d/B%d)*100)' % (row, row, row, row)
            cellObj.value = n

    for row, cellObj in enumerate(list(ws2.columns)[5], start=1):
        if row == 1:
            continue
        elif row == 2:
            continue
        else:
            n = """=IF(E%d<=40,"Excellent",IF(E%d<=60,"Good",IF(E%d<=80,"Fair","Poor")))""" % (row, row, row)
            cellObj.value = n

    for row, cellObj in enumerate(list(ws2.columns)[9], start=1):
        if row == 1:
            continue
        elif row == 2:
            continue
        else:
            n = """=IF((VALUE(I%d)*100)<=40,"Excellent",IF((VALUE(I%d)*100)<=60,"Good",IF((VALUE(I%d)*100)<=80,"Fair","Poor")))""" % (
            row, row, row)
            cellObj.value = n

def SheetBuffer():
    # Sheet Buffer
    dicthost = dict(HostList=hostname)
    dictSmall = OrderedDict([('SmallHits', smallhits), ('SmallMisses', smallmiss)])
    dictMiddle = OrderedDict([('MiddleHits', middlehits), ('MiddleMisses', middlemiss)])
    dictBig = OrderedDict([('BigHits', bighits), ('BigMisses', bigmiss)])
    dictVeryBig = OrderedDict([('VeryBigHits', verybighits), ('VeryBigMisses', verybigmiss)])
    dictLarge = OrderedDict([('LargeHits', largehits), ('LargeMisses', largemiss)])
    dictHuge = OrderedDict([('HugeHits', hugehits), ('HugeMisses', hugemiss)])

    dictCollectedBuff = OrderedDict([('Small', dictSmall), ('Middle', dictMiddle),
                                     ('Big', dictBig), ('VeryBig', dictVeryBig), ('Large', dictLarge),
                                     ('Huge', dictHuge)])

    row = ws3.max_row + 1
    col = 1
    i = 0

    for key in dicthost.keys():
        for item in dicthost[key]:
            ws3.cell(row=row, column=col).value = item
            col += 1

    for key in dictCollectedBuff.keys():
        col = 2
        ws3.cell(row=row + i, column=col).value = key
        col += 1
        for key2 in dictCollectedBuff[key]:
            for item in dictCollectedBuff[key][key2]:
                ws3.cell(row=row + i, column=col).value = int(item)
                col += 1
        i += 1

    for row, cellObj in enumerate(list(ws3.columns)[4], start=1):
        if row == 1:
            continue
        else:
            n = "=C%d+D%d" % (row, row)
            cellObj.value = n

    for row, cellObj in enumerate(list(ws3.columns)[5], start=1):
        if row == 1:
            continue
        else:
            n = "=IF(ISERROR(D%d/E%d),0,(D%d/E%d)*100)" % (row, row, row, row)
            cellObj.value = n

    for row, cellObj in enumerate(list(ws3.columns)[6], start=1):
        if row == 1:
            continue
        else:
            n = """=IF(F%d<=5,"Excellent",IF(F%d<=10,"Good",IF(F%d<=20,"Fair","Poor")))""" % (row, row, row)
            cellObj.value = n

def Compile(path1,path2):
    app = xw.App(visible=False)
    book = app.books.open(path1)
    sys.stdout.flush()
    sleep(10)
    book.save(path2)
    app.kill()
    return openpyxl.load_workbook(path2, data_only=True)

#Call MakeSpreadsheet function
MakeSpreadshet()

# Connect to Devices
wb2 = openpyxl.load_workbook('DeviceList.xlsx', data_only=True)
rs = wb2["DeviceList"]

mylist = []

for cols in rs.iter_cols(min_row=1, min_col=1):
    for cell in cols:
        mylist.append(cell.value)

ip_index = mylist.index("ip")
user_index = mylist.index("username")
pass_index = mylist.index("password")
mylistIP = mylist[ip_index+1:user_index]
mylistUser = mylist[user_index+1:pass_index]
mylistPass = mylist[pass_index+1:]

#loop through devices
for ip, user, passw in zip(mylistIP, mylistUser, mylistPass):
    sys.stdout.flush()
    sleep(0.1)
    print("Connecting to %s" % (ip))
    try:
        net_connect = ConnectHandler(device_type='cisco_ios', ip=ip, username=user, password=passw, timeout=10)
        print("Connected to %s" % (ip))
    except:
        print("Reconecting to %s" % (ip))
        try:
            net_connect = ConnectHandler(device_type='cisco_ios', ip=ip, username=user, password=passw, timeout=10)
            print("Connected to %s" % (ip))
        except:
            print("Reconecting to %s" % (ip))
            try:
                net_connect = ConnectHandler(device_type='cisco_ios', ip=ip, username=user, password=passw, timeout=5)
                print("Connected to %s" % (ip))
            except:
                print("Could not connect to %s\n" % (ip))
                cannotCount = cannotCount + 1
                continue
    # command
    output = net_connect.send_command('show running-config')
    output = output + net_connect.send_command('show inventory')
    output = output + net_connect.send_command('show version')
    output = output + net_connect.send_command('show memory statistics')
    output = output + net_connect.send_command('show processes cpu sorted')
    output = output + net_connect.send_command('show buffers')

    # Parse
    hostname = re.findall(r'^hostname (.*)\s+', output, re.I | re.M)
    PID = re.findall(r'\s*PID:(.*)\s*,\sVID', output, re.I | re.M)
    Description = re.findall(r',\s+DESCR:(.*)\s+', output, re.I | re.M)
    SN = re.findall(r',\s+SN:(.*)\s+', output, re.I | re.M)
    IOS = re.findall(r'^Cisco IOS Software,.*\s+\((.*)\),\s+', output, re.I | re.M)
    Version = re.findall(r'^Cisco IOS Software.*,\s+Version\s+(.*),\s+', output, re.I | re.M)
    Uptime = re.findall(r'\s*uptime is\s+(.*)\s*', output, re.I | re.M)
    DRAM = re.findall(r'^cisco.*processor.*\swith\s(.*)\/', output, re.I | re.M)
    MemTot = re.findall(r'^Processor\s+\w+\s+(\d+)\s+', output, re.I | re.M)
    MemUse = re.findall(r'^Processor\s+\w+\s+\d+\s+(\d+)\s+', output, re.I | re.M)
    MemFree = re.findall(r'^Processor\s+\w+\s+\d+\s+\d+\s+(\d+)\s+', output, re.I | re.M)
    Cpu5sec = re.findall(r'\sfive\sseconds:\s(.*);\sone', output, re.I | re.M)
    Cpu1min = re.findall(r'\sone\sminute:\s(.*);\sfive', output, re.I | re.M)
    Cpu5min = re.findall(r'\sfive\sminutes:\s(.*)\s*', output, re.I | re.M)
    smallhits = re.findall(r'^Small.*\s+.*\s+(\d+)\s+hits,', output, re.I | re.M)
    smallmiss = re.findall(r'^Small.*\s+.*\s+.*(\d+)\s+misses,', output, re.I | re.M)
    middlehits = re.findall(r'^Middle.*\s+.*\s+(\d+)\s+hits,', output, re.I | re.M)
    middlemiss = re.findall(r'^Middle.*\s+.*\s+.*(\d+)\s+misses,', output, re.I | re.M)
    bighits = re.findall(r'^Big.*\s+.*\s+(\d+)\s+hits,', output, re.I | re.M)
    bigmiss = re.findall(r'^Big.*\s+.*\s+.*(\d+)\s+misses,', output, re.I | re.M)
    verybighits = re.findall(r'^VeryBig.*\s+.*\s+(\d+)\s+hits,', output, re.I | re.M)
    verybigmiss = re.findall(r'^VeryBig.*\s+.*\s+.*(\d+)\s+misses,', output, re.I | re.M)
    largehits = re.findall(r'^Large.*\s+.*\s+(\d+)\s+hits,', output, re.I | re.M)
    largemiss = re.findall(r'^Large.*\s+.*\s+.*(\d+)\s+misses,', output, re.I | re.M)
    hugehits = re.findall(r'^Huge.*\s+.*\s+(\d+)\s+hits,', output, re.I | re.M)
    hugemiss = re.findall(r'^Huge.*\s+.*\s+.*(\d+)\s+misses,', output, re.I | re.M)

    # Call function SheetDevice
    SheetDevice()

    # Call function SheetMemCPU
    SheetMemCPU()

    # Call function SheetBuffer
    SheetBuffer()

    print("Data with Hostname %s have been inputed\n" % (hostname))

wb.save('OutputData1.xlsx')
sys.stdout.flush()
sleep(5)

df1 = Compile("OutputData1.xlsx","OutputData2.xlsx")
rs2 = df1["Mem_CPU"]
rs3 = df1["Buffer"]
rs4 = df1["Summary"]
countMaxRow1 = rs2.max_row
countMaxRow2 = rs3.max_row


# #Summary
listMemCpu = []
listBuffer = []

for row in rs2.iter_rows(min_col=1, max_col=10, min_row=3, max_row=countMaxRow1):
    for index, cell in enumerate(row):
        if index == 0 and  cell.value != None:
            listMemCpu.append(cell.value)
        elif index == 5 and cell.value != None:
            listMemCpu.append(cell.value)
        elif index == 9 and cell.value != None:
            listMemCpu.append(cell.value)

for row in rs3.iter_rows(min_col=1, max_col=7, min_row=2, max_row=countMaxRow2):
    for index, cell in enumerate(row):
        if index == 0 and cell.value != None:
            listBuffer.append(cell.value)
        elif index == 6 and  cell.value != None:
            listBuffer.append(cell.value)

r = 2
c = 1
for j, val in enumerate(listMemCpu, start=1):
    rs4.cell(row=r, column=c).value = val
    c += 1
    if j % 3 == 0:
        r += 1
        c = 1

# r2 = 2
# c2 = rs4.max_column
# for j, val in enumerate(listBuffer, start=1):
#     rs4.cell(row=r2, column=c2).value = val
#     c += 1
#     if j % 3 == 0:
#         r += 1
#         c = 1


# make Chart
chart = BarChart3D()
chart.title = "Memory Utilization Chart"
chart.x_axis.title = "Hostname"
chart.y_axis.title = "Utilization %"
chart.y_axis.scaling.max = 60.000

data = Reference(rs2, min_col=5, min_row=3, max_col=5, max_row=countMaxRow1)
titleObj = Reference(rs2, min_col=1, min_row=3, max_col=1, max_row=countMaxRow1)
chart.add_data(data, titles_from_data=False)
chart.set_categories(titleObj)
chart.legend = None

rs2.add_chart(chart, "L5")
df1.save("PreventiveMaintenance.xlsx")

os.remove("OutputData1.xlsx")
os.remove("OutputData2.xlsx")

print("")
print("The number of device that can't connect: " + str(cannotCount))
print('Done')
input("Press Enter to close...")
sys.exit()


